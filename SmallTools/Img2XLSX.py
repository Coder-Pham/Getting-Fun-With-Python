import time

from openpyxl import Workbook
from openpyxl.styles import colors, fills
from PIL import Image


def hexcode(r, g, b):
    '''Hexcode

    Arguments:
        r {int} -- 0-255 of RED
        g {int} -- 0-255 of GREEN
        b {int} -- 0-255 of BLUE

    Returns:
        string -- Hex code of color with FF-prefix
    '''

    return 'FF%02X%02X%02X' % (r, g, b)


def main():
    File = input("Enter Image file name: ")
    print('Processing Image ...')
    count = time.time()

    # Process Image in each pixel
    img = Image.open(File)
    img = img.rotate(0, expand=1)
    width, height = img.size

    ratio = width / height

    # For some god damm reason, Excel can only display file below 1400kb without error
    # I can only display on 310Col X 430Row in 50% magnification
    # So image's width <= 310px
    # image's height <= 430px

    # Resize img
    if (ratio < 1):
        # width < height => Vertical image
        resize_rate = height / 430
        height = int(height // resize_rate)
        width = int(ratio*height)
    else:
        # height < width => Horizontal image
        resize_rate = width / 353
        width = int(width // resize_rate)
        height = int(width // ratio)

    img = img.resize((width, height), Image.BICUBIC)

    # EXCEL
    wb = Workbook()
    sheet = wb.active
    for x in range(width):
        for y in range(height):
            sheet.row_dimensions[y + 2].height = 1
            r, g, b = img.getpixel((x, y))
            # Fill in 1 cell
            color = hexcode(r, g, b)
            sheet.cell(row=y + 2, column=x + 1).fill = fills.PatternFill(
                fill_type='solid', start_color=color, end_color=color)

    for col in sheet.columns:
        column = col[0].column
        sheet.column_dimensions[column].width = 0.15

    wb.save('result.xlsx')
    print('Finish in {:0.5f}s'.format(time.time() - count))


if __name__ == '__main__':
    main()
